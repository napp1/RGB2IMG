from PIL import Image
import openpyxl

#apre file xlsx
file_xlsx = openpyxl.load_workbook('file.xlsx')
sheet = file_xlsx['Foglio1']

#legge file xlsx e per ogni riga colore e crea l'immagine
#il range dipende dal numero di righe del foglio di calcolo
for row_position in range(2, 203):
    codice=sheet.cell(row=row_position, column=1).value
    color_r=int(sheet.cell(row=row_position, column=2).value)
    color_g=int(sheet.cell(row=row_position, column=3).value)
    color_b=int(sheet.cell(row=row_position, column=4).value)
    print(codice, color_r, color_g, color_b)
    img = Image.new('RGB', (1000, 1000), color = (color_r, color_g, color_b))
    img.save(codice+'.png')
